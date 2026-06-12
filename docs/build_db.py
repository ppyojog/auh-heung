# -*- coding: utf-8 -*-
"""
어흥 : 야수의 생존 — 마스터 DB 생성기 (정본 데이터 = 이 스크립트)
실행: python build_db.py  →  AuhHeung_DB.xlsx 생성
- 데이터를 여기서 수정하면 git diff로 변경점이 깔끔히 보이고, 엑셀로 보기 좋게 출력됨.
- 코드(../lib/main.dart) 밸런스와 동기화 유지.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HDR_FILL = PatternFill("solid", fgColor="3A2F28")
HDR_FONT = Font(bold=True, color="F3D89A", size=11)
TITLE_FONT = Font(bold=True, size=14, color="E8A33D")
KEY_FONT = Font(bold=True, color="F3D89A")
THIN = Side(style="thin", color="55463A")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

def sheet(name, headers, rows, note=None, widths=None):
    ws = wb.create_sheet(name)
    r = 1
    if note:
        c = ws.cell(r, 1, note); c.font = Font(italic=True, color="9C8C7E"); r += 2
    for ci, h in enumerate(headers, 1):
        c = ws.cell(r, ci, h); c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center"); c.border = BORDER
    hdr_row = r; r += 1
    for row in rows:
        for ci, v in enumerate(row, 1):
            c = ws.cell(r, ci, v); c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=isinstance(v, str) and len(str(v)) > 30)
        r += 1
    # 열 너비
    for ci in range(1, len(headers) + 1):
        if widths and ci - 1 < len(widths):
            w = widths[ci - 1]
        else:
            w = max([len(str(headers[ci - 1]))] + [len(str(row[ci - 1])) for row in rows if ci - 1 < len(row)] + [8])
        ws.column_dimensions[chr(64 + ci)].width = min(max(w + 2, 9), 64)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    return ws

# ── Overview ──
ov = wb.active; ov.title = "Overview"
ov.cell(1, 1, "어흥 : 야수의 생존 — Master DB").font = TITLE_FONT
ov_rows = [
    ("게임명", "어흥 : 야수의 생존 (AuhHeung : Beast Survivor)"),
    ("장르", "뱀파이어 서바이버즈류 / bullet heaven / 로그라이트"),
    ("USP", "약자→최강 포식자 권력 판타지 + 충신 herald 뽕 + 인디 정체성"),
    ("플랫폼", "Flutter Web(단일파일) → 추후 모바일 포팅"),
    ("그래픽/사운드", "에셋 0 — CustomPainter 네온 / PCM 합성"),
    ("빌드", "Flutter 3.24.5, --web-renderer html, withOpacity만"),
    ("라이브", "https://ppyojog.github.io/auh-heung/"),
    ("저장소", "github.com/ppyojog/auh-heung"),
    ("정본 문서", "docs/GDD.md (설계) + 이 DB(데이터). 생성: docs/build_db.py"),
    ("버전관리", "git 히스토리. 파일명에 버전 숫자 X."),
]
for i, (k, v) in enumerate(ov_rows, start=3):
    a = ov.cell(i, 1, k); a.font = KEY_FONT; a.fill = HDR_FILL; a.border = BORDER
    b = ov.cell(i, 2, v); b.border = BORDER
ov.column_dimensions['A'].width = 16; ov.column_dimensions['B'].width = 66
ov.freeze_panes = "A3"

sheet("Characters",
      ["ID", "이름", "아이콘", "시작무기", "공격", "체력", "이동", "콘셉트"],
      [("white", "백호 부족장", "🐯", "발톱", 1.0, 1.0, 1.0, "균형"),
       ("black", "그림자 흑표", "🐆", "송곳니", 1.3, 0.7, 1.12, "유리대포(고위험)"),
       ("iron", "무쇠뿔 들소", "🐃", "포효", 0.9, 1.6, 0.82, "탱크(저속·고체력)")],
      widths=[8, 14, 6, 9, 6, 6, 6, 20])

sheet("Weapons",
      ["ID", "이름", "유형", "쿨다운", "데미지", "핵심 스케일", "최대Lv"],
      [("claw", "발톱 폭풍", "투사체", "0.8×0.9^(Lv-1)/fireMult", "9+Lv×4", "투사체 1(+@2,4,6)·관통@5", 8),
       ("fang", "회전 송곳니", "궤도 DPS", "상시", "DPS 22+Lv×14", "개수=Lv·반경60+Lv×4·접촉13", 6),
       ("roar", "포효", "즉발 광역", "2.6×0.93^(Lv-1)/fireMult", "8+Lv×6", "반경70+Lv×16·넉백14", 7),
       ("bolt", "벼락", "연쇄 번개", "1.5×0.9^(Lv-1)/fireMult", "10+Lv×6", "연쇄1+Lv·사정210·×0.82", 7),
       ("spike", "가시밭", "광역 지대", "1.8×0.92^(Lv-1)/fireMult", "8+Lv×5", "반경42+Lv×7·오프셋±160", 7)],
      note="데미지엔 dmgMult(야성·캐릭터·광폭화), 쿨다운엔 fireMult(분노·광폭화) 적용",
      widths=[8, 12, 10, 26, 12, 30, 7])

sheet("Evolutions",
      ["진화무기", "원본", "조건", "효과"],
      [("천 개의 발톱", "claw", "발톱 Lv8 + 분노 Lv3", "16발 전방위·데미지×1.7·관통3·쿨0.5"),
       ("죽음의 고리", "fang", "송곳니 Lv6 + 야성 Lv3", "개수+3·반경×1.4·DPS×1.9·접촉17"),
       ("대지진", "roar", "포효 Lv7 + 가죽 Lv3", "반경×1.8·데미지×2.2·넉백30")],
      note="레벨업 시 조건 충족하면 진화 선택지 우선 노출 (뱀서의 핵심 훅)",
      widths=[14, 8, 22, 36])

sheet("Passives",
      ["ID", "이름", "레벨당 효과"],
      [("wild", "야성", "공격력 +12%"), ("hide", "가죽", "최대체력 +25, 즉시 25 회복"),
       ("wind", "바람", "이동속도 +10%"), ("hunger", "굶주림", "수집범위 +16"),
       ("rage", "분노", "공속 +10%")],
      widths=[10, 10, 28])

sheet("Enemies",
      ["타입", "HP", "이동속도", "접촉피해", "반경", "등장", "드랍"],
      [("grunt 잡몹", "(9+0.5t)×diff", "44+0.12t", "(4.5+0.03t)×diff", 11, "항상", "구슬1"),
       ("fast 쾌속", "(9+0.5t)×0.65×diff", "72+0.17t", "(4+0.03t)×diff", 9, "t>42, 32%", "구슬1"),
       ("tank 탱크", "(9+0.5t)×3.0×diff", "30+0.05t", "(8.5+0.05t)×diff", 18, "t>78, 13%", "구슬3"),
       ("boss 보스", "(240+6t)×diff", "40", "(22+0.08t)×diff", 30, "90초마다", "구슬14")],
      note="t=생존 경과 초, diff=동적 난이도 배수. 스폰: 시작1.4s·간격 max(0.5,2.0-0.011t)·1회 1+(t÷48)·최대120체",
      widths=[12, 20, 12, 20, 6, 12, 8])

sheet("Ultimate_광기",
      ["항목", "값"],
      [("게이지 최대", 75), ("충전", "잡몹+1 / 탱크+3 / 보스+14 (×diff)"),
       ("해방 피해", "60+레벨×12 (전 적)"), ("넉백", 70),
       ("광폭화", "5초 · 공속×1.6 · 공격×1.35 · 이동×1.18")],
      widths=[14, 44])

sheet("DynamicDifficulty",
      ["상황", "발동 조건", "선택 → 효과", "타이니 반응"],
      [("후퇴 제안", "HP<32% & diff>0.65", "물러난다 → diff−0.4·체력35%회복·적정리", "호랑이는 물러설 때를 아는 법이죠"),
       ("(후퇴 거절)", "—", "계속 사냥 → 광기 +30%", "역시 호랑이답습니다! 물러섬을 모르는 분!"),
       ("진격 제안", "HP>60% & ~55초마다 & diff<2.0", "쳐들어간다 → diff+0.45", "이래야 맛이 나죠! 방심하면 한 입에 끝"),
       ("(진격 거절)", "—", "천천히 간다 → 변동 없음", "신중함도 맹수의 덕목이죠")],
      note="diff 0.6~2.2(시작 1.0). 높을수록 적 강함 + XP·광기 획득 ×diff(빠른 성장). 선택 후 32초 쿨다운.",
      widths=[12, 26, 32, 34])

sheet("Herald_충신",
      ["트리거", "발동 시점", "강제표시", "예시 대사"],
      [("greet", "런 시작", "O", "대장님, 오늘도 의회 놈들의 간담을 서늘하게 해주시죠."),
       ("level", "레벨업", "O", "또 강해지셨습니까?! 누가 감히 대장님을 양이라 했습니까!"),
       ("boss", "보스 격파", "O", "보셨습니까! 의회의 거수가 대장님 발톱에 찢겼습니다!"),
       ("ult", "어흥! 궁극", "O", "크하핫—! 대륙이 대장님 포효 앞에 무릎 꿇습니다!!"),
       ("low", "HP<25%(9초쿨)", "X", "이건 일부러 봐주시는 거죠…? 그렇죠?!"),
       ("streak", "30킬마다", "X", "대장님 지나간 자리엔 시체만 쌓입니다!"),
       ("milestone", "1·2·3분 생존", "O", "3분 생존… 대장님은 이미 살아있는 전설이십니다!")],
      note="전부 텍스트(무음) → 회사에서 몰래 해도 소리 없이 뽕. 표시 2.8초·쿨 0.8초.",
      widths=[12, 16, 8, 46])

sheet("Meta_영구강화",
      ["ID", "아이콘", "이름", "효과(레벨당)", "기본가", "최대Lv", "비용증가"],
      [("atk", "💪", "맹수의 발톱", "시작 공격력 +5%", 30, 10, "×1.55"),
       ("hp", "❤", "두꺼운 가죽", "시작 체력 +15", 25, 10, "×1.55"),
       ("spd", "🌬", "바람의 다리", "시작 이동속도 +4%", 30, 6, "×1.55"),
       ("pick", "🧲", "굶주린 코", "수집 범위 +8", 25, 6, "×1.55"),
       ("gain", "🦷", "전리품 사냥꾼", "송곳니 획득 +12%", 40, 8, "×1.55")],
      note="화폐=송곳니. 사망 시 적립: (킬+생존초+레벨×8)×(1+0.12×gain Lv). 비용=기본가×1.55^Lv. localStorage 영구 저장.",
      widths=[8, 6, 14, 22, 8, 8, 10])

sheet("BalanceConstants",
      ["상수", "값", "설명"],
      [("baseMaxHp", 120, "시작 최대 체력"), ("baseSpeed", 156, "기본 이동속도(px/s)"),
       ("pickupRange", 58, "구슬 수집 기본 범위"), ("xpStart", 4, "레벨1 요구 XP"),
       ("xpGrowth", "prev×1.26+2", "다음 레벨 요구 XP"), ("enemyCap", 120, "동시 최대 적"),
       ("bossInterval", 90, "보스 주기(초)"), ("rageMax", 75, "광기 게이지 최대"),
       ("diffRange", "0.6~2.2", "동적 난이도 배수 범위")],
      widths=[14, 14, 26])

sheet("Roadmap",
      ["상태", "항목"],
      [("완료", "네온 아트 + 손맛(데미지숫자·흔들림·파티클)"),
       ("완료", "코드 합성 사운드(효과음7 + 음소거)"),
       ("완료", "무기 진화 + 신무기(벼락·가시밭)"),
       ("완료", "캐릭터 3종"),
       ("완료", "초반 난이도 완화"),
       ("완료", "조작 개선 + 어흥! 궁극 + 스토리 도입"),
       ("완료", "충신 herald 시스템(무음 텍스트)"),
       ("완료", "타이니 동적 난이도 선택 대화"),
       ("완료", "메타 영구강화(송곳니 화폐 → 전리품 상점) + 위협도 표시"),
       ("완료", "무해·귀여움 컨셉: 캐릭터 절차 애니메이션(숨쉬기·통통·눈깜빡·꼬리)+캐릭터별 외형, 타이니 등장연출"),
       ("다음★", "귀여운 코스메틱 스킨(윤리적 수익화) / 포식(Devour)"),
       ("다음", "무기/캐릭터 해금(송곳니) / 보스 패턴·서사"),
       ("다음", "CanvasKit(스웜300+) / 엔드리스·데일리·리더보드")],
      widths=[8, 50])

sheet("ChangeLog",
      ["단계", "내용"],
      [("legacy", "투자 은유 카드게임(폐기, 로컬 v28.0 아카이브)"),
       ("피벗1", "텍스트 선택지 → Reigns식 스와이프 카드"),
       ("피벗2", "뱀파이어 서바이버즈류(현재)"),
       ("이후", "git 커밋 로그로 추적")],
      widths=[10, 56])

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "AuhHeung_DB.xlsx")
wb.save(out)
print("saved:", out, os.path.getsize(out), "bytes,", len(wb.sheetnames), "sheets")
